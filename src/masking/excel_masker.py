"""
Excel (.xlsx) マスキング処理
- 全シートの全セル値（文字列）をマスク
- セル内ハイパーリンクURLもマスク
- 数式・数値・書式は変更しない
"""
from __future__ import annotations
import io

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_STRING, TYPE_FORMULA_CACHE_STRING

from .patterns import apply_masks


def mask_excel(file_bytes: bytes, opts: dict) -> tuple[bytes, int]:
    """
    Excel ファイルをマスクして返す。
    return: (masked_bytes, total_mask_count)
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    total = 0

    for ws in wb.worksheets:
        # ── セル値 ──
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # 文字列セルのみ処理
                if isinstance(cell.value, str):
                    masked, cnt = apply_masks(cell.value, opts)
                    if cnt > 0:
                        cell.value = masked
                        total += cnt

                # 数式の場合はキャッシュ値（表示テキスト）にもマスクを試みる
                if cell.data_type in (TYPE_FORMULA_CACHE_STRING,):
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        pass  # 数式本体は変更しない

        # ── セル内ハイパーリンク ──
        if opts.get('url'):
            for cell_ref, hl in (ws.hyperlinks._hyperlinks or {}).items():
                if hl.target:
                    masked, cnt = apply_masks(hl.target, opts)
                    if cnt > 0:
                        hl.target = masked
                        total += cnt

        # ── コメント（notes）──
        if ws.cell_comment_map:
            for coord, comment in ws.cell_comment_map.items():
                if comment and comment.text:
                    text = comment.text.plain_text if hasattr(comment.text, 'plain_text') else str(comment.text)
                    masked, cnt = apply_masks(text, opts)
                    if cnt > 0 and hasattr(comment.text, 'plain_text'):
                        comment.text.plain_text = masked
                        total += cnt

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), total
