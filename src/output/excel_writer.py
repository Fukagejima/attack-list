"""
Excel / CSV 出力層
- アタックリストシート
- 業界ヒートマップシート
- メタ情報シート
"""
import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from config.settings import OUTPUT_DIR
from utils.logger import get_logger

logger = get_logger(__name__)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# セルスタイル定数
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
BORDER_SIDE = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE
)


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len * 1.8 + 2, 50)


def write_attack_list_excel(
    companies: list[dict],
    industry_scores: list[dict],
    meta: dict,
) -> str:
    """
    Excelファイルを生成してパスを返す。
    - Sheet1: アタックリスト
    - Sheet2: 業界ヒートマップ
    - Sheet3: メタ情報
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"アタックリスト_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    _write_attack_sheet(wb, companies)
    _write_heatmap_sheet(wb, industry_scores)
    _write_meta_sheet(wb, meta)

    # デフォルトシート削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    logger.info(f"Excel出力完了: {filepath}")
    return filepath


def write_csv(companies: list[dict]) -> str:
    """CSVのみ出力"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"アタックリスト_{timestamp}.csv"
    filepath = os.path.join(OUTPUT_DIR, filename)

    df = pd.DataFrame(companies)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    logger.info(f"CSV出力完了: {filepath}")
    return filepath


def _write_attack_sheet(wb: openpyxl.Workbook, companies: list[dict]):
    ws = wb.create_sheet("📋 アタックリスト")
    ws.freeze_panes = "A2"

    columns = [
        ("優先順位", "rank"),
        ("企業名", "company_name"),
        ("業界", "industry"),
        ("売上規模", "revenue_scale"),
        ("従業員規模", "employee_scale"),
        ("ICP適合度\n(1-10)", "icp_fit_score"),
        ("ソリューション\nフィット(1-10)", "solution_fit_score"),
        ("競合飽和度\n(1-10)", "competitor_saturation_score"),
        ("総合スコア\n(100点満点)", "total_score"),
        ("自社との相性", "compatibility_comment"),
        ("初回アプローチヒント", "approach_hint"),
        ("情報源", "public_info_source"),
        ("データ確度", "data_confidence"),
    ]

    # ヘッダー
    for col_idx, (header, _) in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(columns))
    ws.row_dimensions[1].height = 30

    # データ行
    for row_idx, company in enumerate(companies, 2):
        for col_idx, (_, key) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=company.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 総合スコア列に条件付き書式（カラースケール）
    score_col = get_column_letter(9)
    last_row = len(companies) + 1
    ws.conditional_formatting.add(
        f"{score_col}2:{score_col}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )

    _auto_column_width(ws)


def _write_heatmap_sheet(wb: openpyxl.Workbook, industry_scores: list[dict]):
    ws = wb.create_sheet("🌡️ 業界ヒートマップ")
    ws.freeze_panes = "B2"

    # スコアキー定義
    score_keys = [
        ("市場規模", "market_size"),
        ("実績伸び率", "growth_actual"),
        ("将来伸び率", "growth_future"),
        ("市場ニーズ", "market_overview"),
        ("競合少なさ", "competitor_count_score"),
        ("競合弱さ", "competitor_strength_score"),
        ("市場余地", "market_gap_score"),
        ("解決策適合", "solution_match"),
        ("実績親和性", "past_case_affinity"),
        ("リソース適合", "resource_fit"),
        ("総合スコア", "total_score"),
    ]

    # ヘッダー行
    ws.cell(row=1, column=1, value="業界名")
    for col_idx, (label, _) in enumerate(score_keys, 2):
        ws.cell(row=1, column=col_idx, value=label)
    ws.cell(row=1, column=len(score_keys) + 2, value="サマリー")
    ws.cell(row=1, column=len(score_keys) + 3, value="リスク")
    _style_header_row(ws, 1, len(score_keys) + 3)
    ws.row_dimensions[1].height = 25

    for row_idx, ind in enumerate(industry_scores, 2):
        ws.cell(row=row_idx, column=1, value=ind.get("industry", ""))
        scores = ind.get("scores", {})
        for col_idx, (_, key) in enumerate(score_keys, 2):
            val = ind.get("total_score") if key == "total_score" else scores.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        ws.cell(row=row_idx, column=len(score_keys) + 2, value=ind.get("summary", "")).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=len(score_keys) + 3, value=ind.get("risks", "")).alignment = Alignment(wrap_text=True)

    # カラースケール（スコア列すべてに適用）
    last_row = len(industry_scores) + 1
    for col_idx in range(2, len(score_keys) + 2):
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    _auto_column_width(ws)


def _write_meta_sheet(wb: openpyxl.Workbook, meta: dict):
    ws = wb.create_sheet("ℹ️ 分析メタ情報")
    rows = [
        ("生成日時", meta.get("generated_at", "")),
        ("会社名", meta.get("company_name", "")),
        ("サービス名", meta.get("service_name", "")),
        ("強み", meta.get("strengths", "")),
        ("既存取引先業界", meta.get("existing_industry", "")),
        ("分析対象業界数", meta.get("target_industry_count", "")),
        ("出力企業数", meta.get("company_count", "")),
        ("使用モデル", meta.get("model", "")),
        ("注意事項", "本リストはAIが公開情報を元に生成した推定情報です。最終判断は必ず人間が行ってください。"),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v)).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
